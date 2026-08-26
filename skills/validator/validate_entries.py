# -*- coding: utf-8 -*-
"""
validate_entries.py — 《数据处理与分析》学生数据批量校验脚本 v1.1
用法: python validate_entries.py 学生文件.xlsx [--master 规范清单母表.xlsx] [-o 报告.xlsx]
读取学生录入模板(v2.0/v2.1 的 28 列结构),独立重算全部规则(不依赖表内公式),
输出逐条问题清单与摘要报告。规则编号与模板【字段定义】对应。
"""
import argparse, re, sys, datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

FONT = "微软雅黑"
COLS = ["条目ID","文件名称","文号","位阶","制定区域","发布机关","公布日期","施行日期",
        "效力状态","失效或修订日期","主题标签主","主题标签次","条文定位",
        "规范主体","适用条件","行为模式类型","行为模式内容","法律后果",
        "原文摘录","关联条目","央地权限配置","地方自主权类型","核实来源",
        "录入人","版本","录入日期","备注","自检结果"]
HIERARCHY = {"法律","行政法规","部门规章","地方立法","规范性文件","政策文件"}
TOPICS = {"规划与区划","海洋空间与用海管理","海洋可再生能源","海洋油气与矿产资源",
          "能源基础设施","电力基础设施","价格、财税与市场机制","生态环境与海岛保护",
          "海洋权益与涉外事务","海洋观测与数据统计"}
NORM_TYPE = {"义务性","授权性","禁止性","定义性","程序性","指导性","惩罚性"}
CL_CONFIG = {"统一执行","授权细化","因地制宜","未作规定","中央专属"}
LOCAL_AUT = {"执行性","细化性","创设性","可协商性","无权限(存疑)"}
AI_NAMES = ["豆包","deepseek","chatgpt","claude","kimi","文心","通义","元宝","gemini","copilot"]
CN_NUM = "一二三四五六七八九十百零"

def norm_name(s):
    return re.sub(r"[《》\s]", "", str(s or ""))

def is_date(v):
    return isinstance(v, (datetime.date, datetime.datetime))

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("student")
    ap.add_argument("--master", default=None)
    ap.add_argument("-o", "--out", default=None)
    a = ap.parse_args()

    swb = load_workbook(a.student, data_only=True)
    if "数据录入" not in swb.sheetnames:
        sys.exit("找不到【数据录入】工作表")
    ws = swb["数据录入"]

    master = {}
    if a.master:
        mwb = load_workbook(a.master, data_only=True)
        mws = mwb["规范清单"]
        for r in range(5, 205):
            nm = mws.cell(row=r, column=2).value
            if nm:
                master[norm_name(nm)] = {
                    "位阶": mws.cell(row=r, column=4).value,
                    "公布": mws.cell(row=r, column=7).value,
                    "施行": mws.cell(row=r, column=8).value,
                    "状态": mws.cell(row=r, column=9).value,
                }

    rows = []
    for r in range(8, 208):
        vals = [ws.cell(row=r, column=j).value for j in range(1, 29)]
        content = [v for i, v in enumerate(vals) if i not in (0, 23, 27) and v not in (None, "")]
        if content:
            rows.append((r, vals))

    issues = []  # (行号, 条目ID, 级别, 检查项, 详情)
    def add(r, eid, lvl, item, detail):
        issues.append((r, eid or "", lvl, item, detail))

    ids, locs = {}, {}
    hier_seen, arabic, chinese = set(), 0, 0

    for r, v in rows:
        eid = str(v[0] or "")
        d = dict(zip(COLS, v))
        # ① 条目ID
        if not eid:
            add(r, eid, "错误", "①条目ID", "缺条目ID")
        else:
            ids.setdefault(eid, []).append(r)
            if not re.fullmatch(r"S\d{2}-\d{3}", eid):
                add(r, eid, "提醒", "①条目ID", f"格式非标准(应为 S##-###,如 S24-015);当前「{eid}」。v2.1 模板已改为自动生成,可消除此类问题")
        # 必填
        for k, tag in [("文件名称","②"),("位阶","④"),("制定区域","⑤"),("主题标签主","⑪"),
                       ("规范主体","⑭"),("适用条件","⑮"),("行为模式类型","⑯"),
                       ("行为模式内容","⑰"),("原文摘录","⑲"),("核实来源","㉓"),
                       ("录入人","㉔"),("版本","㉕")]:
            if not d[k]:
                add(r, eid, "错误", f"{tag}{k}", "必填项为空")
        # 枚举
        if d["位阶"] and d["位阶"] not in HIERARCHY:
            add(r, eid, "错误", "④位阶", f"取值「{d['位阶']}」不在受控词表")
        if d["位阶"]: hier_seen.add(d["位阶"])
        if d["主题标签主"] and d["主题标签主"] not in TOPICS:
            add(r, eid, "错误", "⑪主题标签", f"取值「{d['主题标签主']}」不在受控词表")
        if d["主题标签次"] and d["主题标签次"] == d["主题标签主"]:
            add(r, eid, "错误", "⑫主题标签", "次标签与主标签相同")
        if d["行为模式类型"] and d["行为模式类型"] not in NORM_TYPE:
            add(r, eid, "错误", "⑯行为模式类型", f"取值「{d['行为模式类型']}」不在受控词表")
        # 日期
        if d["施行日期"] and not is_date(d["施行日期"]):
            add(r, eid, "错误", "⑧施行日期", f"非日期格式:「{d['施行日期']}」")
        if d["效力状态"] and d["效力状态"] != "现行有效" and not d["失效或修订日期"]:
            add(r, eid, "错误", "⑩失效日期", f"状态为「{d['效力状态']}」但未填失效/修订日期")
        if is_date(d["失效或修订日期"]) and is_date(d["施行日期"]) and d["失效或修订日期"] <= d["施行日期"]:
            add(r, eid, "错误", "⑩失效日期", "失效日期不晚于施行日期")
        # ⑬ 定位
        loc = str(d["条文定位"] or "")
        if loc:
            if "条" not in loc and "段" not in loc and "项" not in loc:
                add(r, eid, "错误", "⑬条文定位", f"未定位到条/款/项或段:「{loc}」")
            if re.search(r"第[0-9]+条", loc): arabic += 1
            if re.search(f"第[{CN_NUM}]+条", loc): chinese += 1
            key = (norm_name(d["文件名称"]), re.sub(r"\s", "", loc))
            locs.setdefault(key, []).append((r, eid, str(d["行为模式内容"] or "")))
        # 字数
        if d["行为模式内容"] and len(str(d["行为模式内容"])) > 120:
            add(r, eid, "错误", "⑰行为模式内容", f"超120字({len(str(d['行为模式内容']))}字)")
        if d["原文摘录"] and len(str(d["原文摘录"])) > 200:
            add(r, eid, "错误", "⑲原文摘录", f"超200字({len(str(d['原文摘录']))}字)")
        # 疑似复制
        if d["行为模式内容"] and d["原文摘录"]:
            sim = SequenceMatcher(None, str(d["行为模式内容"]), str(d["原文摘录"])).ratio()
            if sim >= 0.8:
                add(r, eid, "提醒", "⑰疑似复制", f"行为模式内容与原文摘录相似度{sim:.0%},请改为自己的概括")
        # ㉑㉒ 央地镜像
        if d["制定区域"] == "国家":
            if not d["央地权限配置"]:
                add(r, eid, "错误", "㉑央地权限配置", "国家层条目必填")
            elif d["央地权限配置"] not in CL_CONFIG:
                add(r, eid, "错误", "㉑央地权限配置", f"取值「{d['央地权限配置']}」不在受控词表")
            if d["地方自主权类型"]:
                add(r, eid, "错误", "㉒地方自主权类型", "国家层条目应留空")
        elif d["制定区域"]:
            if not d["地方自主权类型"]:
                add(r, eid, "错误", "㉒地方自主权类型", "地方层条目必填")
            if d["央地权限配置"]:
                add(r, eid, "错误", "㉑央地权限配置", "地方层条目应留空")
        # ㉓ 核实来源
        src = str(d["核实来源"] or "").lower()
        if src and any(x in src for x in AI_NAMES):
            add(r, eid, "错误", "㉓核实来源", f"「{d['核实来源']}」为AI工具,不得作为核实来源(红线,本条计零分)")
        # ㉔ 录入人
        au = str(d["录入人"] or "")
        if au and not re.search(r"\d{6,}", au):
            add(r, eid, "提醒", "㉔录入人", f"「{au}」未含学号;且请勿使用姓名(语料将脱敏开源,统一用「学号 编号」)。v2.1 模板已改为自动生成")
        # 与母表交叉核对
        if master and d["文件名称"]:
            m = master.get(norm_name(d["文件名称"]))
            if m is None:
                add(r, eid, "提醒", "母表核对", f"文件「{d['文件名称']}」不在老师发布的规范清单中(检查是否用了简称或超出认领范围)")
            else:
                if d["位阶"] and m["位阶"] and d["位阶"] != m["位阶"]:
                    add(r, eid, "错误", "母表核对", f"位阶与母表不符:你填「{d['位阶']}」,母表为「{m['位阶']}」")
                if is_date(d["施行日期"]) and is_date(m["施行"]):
                    sd = d["施行日期"]; md = m["施行"]
                    sd = sd.date() if isinstance(sd, datetime.datetime) else sd
                    md = md.date() if isinstance(md, datetime.datetime) else md
                    if sd != md:
                        add(r, eid, "提醒", "母表核对", f"施行日期与母表不符:你填{sd},母表{md}(请回权威库确认,以核实结果为准并写备注)")
                if d["效力状态"] and m["状态"] and d["效力状态"] != m["状态"]:
                    add(r, eid, "提醒", "母表核对", f"效力状态与母表不符:你填「{d['效力状态']}」,母表「{m['状态']}」")

    # 跨行检查
    for eid, rs in ids.items():
        if len(rs) > 1:
            for r in rs:
                add(r, eid, "错误", "①条目ID", f"ID重复,出现在行{rs}")
    for (nm, loc), lst in locs.items():
        if len(lst) > 1:
            dup_rows = set()
            for i in range(len(lst)):
                for j in range(i + 1, len(lst)):
                    sim = SequenceMatcher(None, lst[i][2], lst[j][2]).ratio()
                    if sim >= 0.85:
                        dup_rows.add(lst[i][0]); dup_rows.add(lst[j][0])
            for r, eid, _ in lst:
                if r in dup_rows:
                    add(r, eid, "错误", "重复条款", f"同一文件同一定位、行为模式内容高度相似(行{sorted(dup_rows)}):疑似重复录入:{loc}")
                else:
                    add(r, eid, "提醒", "一文多义拆分", f"同一定位拆分多条(行{[x[0] for x in lst]}):若为一句原文含多种规范,请在备注注明拆分理由:{loc}")
    # 关联条目存在性
    idset = set(ids)
    for r, v in rows:
        rel = str(v[19] or "")
        if rel:
            for token in re.split(r"[,,、\s]+", rel):
                if token and token not in idset:
                    add(r, str(v[0] or ""), "提醒", "⑳关联条目", f"引用的条目「{token}」不存在于本表")
    # 覆盖面与风格
    if len(hier_seen) < 2 and rows:
        add(0, "", "提醒", "覆盖面", f"仅覆盖{len(hier_seen)}个位阶({'、'.join(hier_seen)}),提交要求至少两个——请与组长确认认领包中的其他位阶文件")
    if arabic and chinese:
        add(0, "", "提醒", "定位风格", f"条文定位混用阿拉伯数字({arabic}条)与中文数字({chinese}条),请统一为阿拉伯数字")
    vers = {str(v[24]) for _, v in rows if v[24]}
    if len(vers) > 1:
        add(0, "", "提醒", "版本", f"同一文件内版本号混用:{sorted(vers)}(v1提交应统一为v1)")

    err = sum(1 for i in issues if i[2] == "错误")
    warn = sum(1 for i in issues if i[2] == "提醒")
    bad_rows = {i[0] for i in issues if i[2] == "错误" and i[0]}
    ok = len(rows) - len(bad_rows)
    rate = ok / len(rows) * 100 if rows else 0

    # ---- 报告输出 ----
    out = a.out or a.student.replace(".xlsx", "") + "_校验报告.xlsx"
    rwb = Workbook()
    thin = Side(style="thin", color="BFBFBF"); B = Border(left=thin,right=thin,top=thin,bottom=thin)
    HF = PatternFill("solid", fgColor="1F3864"); HFont = Font(name=FONT,size=9,bold=True,color="FFFFFF")
    s1 = rwb.active; s1.title = "摘要"
    s1["A1"] = "数据校验报告"; s1["A1"].font = Font(name=FONT, size=14, bold=True)
    meta = [("被校验文件", a.student.split("/")[-1]), ("母表", (a.master or "未提供").split("/")[-1]),
            ("校验时间", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("条目总数", len(rows)), ("含错误的条目数", len(bad_rows)),
            ("错误数", err), ("提醒数", warn), ("条目合规率", f"{rate:.1f}%"),
            ("覆盖位阶", "、".join(sorted(hier_seen)) or "—")]
    for i, (k, vv) in enumerate(meta, start=3):
        s1.cell(row=i, column=1, value=k).font = Font(name=FONT, size=10, bold=True)
        s1.cell(row=i, column=2, value=vv).font = Font(name=FONT, size=10)
    s1.column_dimensions["A"].width = 18; s1.column_dimensions["B"].width = 50
    s1.cell(row=14, column=1, value="说明:「错误」须在v2中修正;「提醒」须核实或说明。合规率=不含错误的条目占比。评分以老师发布的评分标准为准。").font = Font(name=FONT, size=9, italic=True, color="808080")

    s2 = rwb.create_sheet("问题清单")
    hdrs = ["行号","条目ID","级别","检查项","详情"]
    for j, h in enumerate(hdrs, start=1):
        c = s2.cell(row=1, column=j, value=h); c.fill = HF; c.font = HFont; c.border = B
        c.alignment = Alignment(horizontal="center")
    for j, w in enumerate([7, 12, 8, 16, 90], start=1):
        s2.column_dimensions[chr(64+j)].width = w
    RED = PatternFill("solid", fgColor="FFC7CE"); YEL = PatternFill("solid", fgColor="FFEB9C")
    for i, (r, eid, lvl, item, det) in enumerate(sorted(issues, key=lambda x:(x[0], x[2])), start=2):
        for j, vv in enumerate([r or "—", eid, lvl, item, det], start=1):
            c = s2.cell(row=i, column=j, value=vv)
            c.font = Font(name=FONT, size=9); c.border = B
            c.alignment = Alignment(wrap_text=True, vertical="top")
        s2.cell(row=i, column=3).fill = RED if lvl == "错误" else YEL
    rwb.save(out)
    print(f"条目{len(rows)} | 错误{err} 提醒{warn} | 合规率{rate:.1f}% | 报告: {out}")

if __name__ == "__main__":
    main()
